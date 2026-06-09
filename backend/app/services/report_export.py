from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

PERIOD_LABELS = {
    "daily": "Diário",
    "weekly": "Semanal",
    "monthly": "Mensal",
}

_TABLE_STYLE = TableStyle(
    [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]
)


def _period_range_label(summary: dict) -> str:
    period_label = PERIOD_LABELS.get(summary["period"], summary["period"])
    start = summary["start_date"].strftime("%d/%m/%Y")
    end = summary["end_date"].strftime("%d/%m/%Y")
    return f"Período: {period_label} ({start} a {end})"


def _summary_rows(summary: dict) -> list[tuple[str, str]]:
    return [
        ("Tarefas concluídas", str(summary["tasks_completed"])),
        ("Tarefas com vencimento no período", str(summary["tasks_due"])),
        ("Taxa de conclusão de tarefas", f"{summary['task_completion_rate']:.1f}%"),
        ("Hábitos ativos", str(summary["habits_active"])),
        ("Check-ins de hábitos", str(summary["habit_checkins"])),
        ("Taxa de conclusão de hábitos", f"{summary['habit_completion_rate']:.1f}%"),
        ("Metas concluídas no período", str(summary["goals_completed"])),
        ("Melhor sequência ativa", f"{summary['best_streak']} dia(s)"),
    ]


def build_pdf(summary: dict, *, user_name: str) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()

    elements = [
        Paragraph("Relatório de Produtividade — DailyFlow", styles["Title"]),
        Paragraph(f"Usuário: {user_name}", styles["Normal"]),
        Paragraph(_period_range_label(summary), styles["Normal"]),
        Spacer(1, 0.6 * cm),
    ]

    summary_table = Table([["Métrica", "Valor"], *_summary_rows(summary)], colWidths=[9 * cm, 6 * cm])
    summary_table.setStyle(_TABLE_STYLE)
    elements.append(summary_table)
    elements.append(Spacer(1, 0.8 * cm))

    elements.append(Paragraph("Detalhamento diário", styles["Heading2"]))
    elements.append(Spacer(1, 0.3 * cm))

    daily_rows = [["Data", "Tarefas concluídas", "Hábitos registrados"]]
    daily_rows += [
        [entry["date"].strftime("%d/%m/%Y"), str(entry["tasks_completed"]), str(entry["habits_completed"])]
        for entry in summary["daily_breakdown"]
    ]
    daily_table = Table(daily_rows, colWidths=[5 * cm, 5 * cm, 5 * cm])
    daily_table.setStyle(_TABLE_STYLE)
    elements.append(daily_table)

    doc.build(elements)
    return buffer.getvalue()


def build_excel(summary: dict, *, user_name: str) -> bytes:
    workbook = Workbook()
    bold = Font(bold=True)

    summary_sheet = workbook.active
    summary_sheet.title = "Resumo"
    summary_sheet.append(["Relatório de Produtividade — DailyFlow"])
    summary_sheet["A1"].font = Font(bold=True, size=14)
    summary_sheet.append([f"Usuário: {user_name}"])
    summary_sheet.append([_period_range_label(summary)])
    summary_sheet.append([])

    summary_sheet.append(["Métrica", "Valor"])
    for cell in summary_sheet[summary_sheet.max_row]:
        cell.font = bold
    for label, value in _summary_rows(summary):
        summary_sheet.append([label, value])

    summary_sheet.column_dimensions["A"].width = 36
    summary_sheet.column_dimensions["B"].width = 16

    daily_sheet = workbook.create_sheet("Detalhamento diário")
    daily_sheet.append(["Data", "Tarefas concluídas", "Hábitos registrados"])
    for cell in daily_sheet[1]:
        cell.font = bold
    for entry in summary["daily_breakdown"]:
        daily_sheet.append([entry["date"].strftime("%d/%m/%Y"), entry["tasks_completed"], entry["habits_completed"]])

    daily_sheet.column_dimensions["A"].width = 16
    daily_sheet.column_dimensions["B"].width = 20
    daily_sheet.column_dimensions["C"].width = 22

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
